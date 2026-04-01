#!/usr/bin/env python3
"""
修复已有产品的成本和价格数据

问题：
- 数据库有 137 个产品
- 只有 16 个有完整成本/价格数据
- 121 个产品缺少成本和价格数据
- 平台费字段为 NULL

任务：
- 为已有产品补充成本数据 (product_costs)
- 为已有产品补充价格数据 (product_prices)
- 计算并填充平台费
"""

import sqlite3
import openpyxl
from datetime import datetime

# 配置
EXCEL_PATH = '/home/wxy/.xiaozhi/files/file_1773984868713_a5025b1a_Foreign_Trade_Management_Template.xlsx.xlsx'
DB_PATH = '/home/wxy/data/ciciERP/data/cicierp.db'
SHEET_INDEX = 2  # 第三个sheet (2026 Product Cost List)
HEADER_ROW = 3
DATA_START_ROW = 4
EXCHANGE_RATE = 6.81
PLATFORM_FEE_RATE = 0.025

# 列索引 (1-based)
COL_PRODUCT_NAME = 1
COL_MODEL = 2
COL_COST_RMB = 3
COL_COST_USD = 4
COL_PROFIT_MARGIN = 5
COL_ALIBABA_FEE = 6
COL_SELLING_PRICE = 7
COL_STOCK_QTY = 8
COL_SIZE = 9
COL_WEIGHT = 10
COL_NOTES = 11
COL_SUPPLIER = 12


def main():
    print("=== ciciERP 产品数据修复工具 ===\n")

    # 1. 加载Excel
    print(f"加载Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb.worksheets[SHEET_INDEX]
    print(f"Sheet: {sheet.title}")

    # 2. 连接数据库
    print(f"连接数据库: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 3. 获取所有产品（id, product_code）
    cursor.execute("SELECT id, product_code FROM products WHERE deleted_at IS NULL")
    products = {row[1]: row[0] for row in cursor.fetchall()}
    print(f"数据库产品数: {len(products)}")

    # 4. 获取已有成本数据的产品
    cursor.execute("SELECT DISTINCT product_id FROM product_costs")
    existing_costs = set(row[0] for row in cursor.fetchall())
    print(f"已有成本记录: {len(existing_costs)}")

    # 5. 获取已有价格数据的产品
    cursor.execute("SELECT DISTINCT product_id FROM product_prices")
    existing_prices = set(row[0] for row in cursor.fetchall())
    print(f"已有价格记录: {len(existing_prices)}")

    # 6. 读取Excel数据
    excel_data = {}
    for row_num in range(DATA_START_ROW, sheet.max_row + 1):
        model = sheet.cell(row=row_num, column=COL_MODEL).value
        if model and str(model).strip():
            model = str(model).strip()
            cost_rmb = sheet.cell(row=row_num, column=COL_COST_RMB).value or 0
            cost_usd = sheet.cell(row=row_num, column=COL_COST_USD).value or 0
            profit_margin = sheet.cell(row=row_num, column=COL_PROFIT_MARGIN).value or 0
            alibaba_fee = sheet.cell(row=row_num, column=COL_ALIBABA_FEE).value
            selling_price = sheet.cell(row=row_num, column=COL_SELLING_PRICE).value
            notes = sheet.cell(row=row_num, column=COL_NOTES).value

            try:
                cost_rmb = float(cost_rmb) if cost_rmb else 0
                cost_usd = float(cost_usd) if cost_usd else 0
                profit_margin = float(profit_margin) if profit_margin else 0
                alibaba_fee = float(alibaba_fee) if alibaba_fee else None
                selling_price = float(selling_price) if selling_price else None
            except (ValueError, TypeError):
                continue

            excel_data[model] = {
                'cost_rmb': cost_rmb,
                'cost_usd': cost_usd,
                'profit_margin': profit_margin,
                'alibaba_fee': alibaba_fee,
                'selling_price': selling_price,
                'notes': notes
            }

    print(f"Excel产品数: {len(excel_data)}\n")

    # 7. 统计
    stats = {
        'cost_added': 0,
        'cost_skipped': 0,
        'price_added': 0,
        'price_skipped': 0,
        'not_found': 0,
        'errors': []
    }

    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

    # 8. 为每个产品补充数据
    for product_code, product_id in products.items():
        if product_code not in excel_data:
            stats['not_found'] += 1
            continue

        data = excel_data[product_code]

        # 跳过没有成本数据的
        if data['cost_rmb'] == 0 and data['cost_usd'] == 0:
            stats['cost_skipped'] += 1
            continue

        try:
            # 补充成本数据
            if product_id not in existing_costs:
                # 计算平台费（基于售价）
                platform_fee = None
                if data['selling_price']:
                    platform_fee = round(data['selling_price'] * PLATFORM_FEE_RATE, 2)

                cursor.execute("""
                    INSERT INTO product_costs (
                        product_id, supplier_id, cost_cny, cost_usd, currency,
                        exchange_rate, profit_margin, platform_fee_rate, platform_fee,
                        sale_price_usd, quantity, purchase_order_id, is_reference,
                        effective_date, notes, created_at, updated_at
                    ) VALUES (?, NULL, ?, ?, 'CNY', ?, ?, 0.025, ?, ?, 1, NULL, 1, NULL, ?, ?, ?)
                """, (
                    product_id,
                    data['cost_rmb'],
                    data['cost_usd'],
                    EXCHANGE_RATE,
                    data['profit_margin'],
                    platform_fee,
                    data['selling_price'],
                    str(data['notes'])[:500] if data['notes'] else None,
                    now, now
                ))
                stats['cost_added'] += 1
                print(f"✓ {product_code}: 添加成本 CNY={data['cost_rmb']}, USD={data['cost_usd']}")

            # 补充价格数据（如果有售价）
            if data['selling_price'] and product_id not in existing_prices:
                sale_price_cny = round(data['selling_price'] * EXCHANGE_RATE, 2)
                platform_fee = round(data['selling_price'] * PLATFORM_FEE_RATE, 2)

                cursor.execute("""
                    INSERT INTO product_prices (
                        product_id, platform, sale_price_cny, sale_price_usd, exchange_rate,
                        profit_margin, platform_fee_rate, platform_fee, is_reference,
                        effective_date, notes, created_at, updated_at
                    ) VALUES (?, 'alibaba', ?, ?, ?, ?, 0.025, ?, 1, NULL, '从成本清单导入', ?, ?)
                """, (
                    product_id,
                    sale_price_cny,
                    data['selling_price'],
                    EXCHANGE_RATE,
                    data['profit_margin'],
                    platform_fee,
                    now, now
                ))
                stats['price_added'] += 1
                print(f"  └─ 添加价格 USD={data['selling_price']}, CNY={sale_price_cny}")

        except Exception as e:
            stats['errors'].append(f"{product_code}: {str(e)}")
            print(f"✗ {product_code}: 错误 - {e}")

    # 9. 提交事务
    conn.commit()

    # 10. 验证结果
    cursor.execute("SELECT COUNT(*) FROM product_costs WHERE cost_cny > 0")
    final_costs = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM product_prices WHERE sale_price_usd > 0")
    final_prices = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM product_costs WHERE platform_fee IS NOT NULL AND platform_fee > 0")
    final_platform_fees = cursor.fetchone()[0]

    # 11. 输出统计
    print("\n=== 修复完成 ===")
    print(f"成本记录添加: {stats['cost_added']}")
    print(f"成本记录跳过(已有): {stats['cost_skipped']}")
    print(f"价格记录添加: {stats['price_added']}")
    print(f"Excel中未找到: {stats['not_found']}")

    if stats['errors']:
        print(f"\n错误数: {len(stats['errors'])}")
        for error in stats['errors'][:10]:
            print(f"  - {error}")

    print(f"\n最终统计:")
    print(f"  成本记录数: {final_costs}")
    print(f"  价格记录数: {final_prices}")
    print(f"  平台费记录数: {final_platform_fees}")

    conn.close()
    wb.close()

    return stats


if __name__ == '__main__':
    main()
