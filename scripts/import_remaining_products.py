#!/usr/bin/env python3
"""
从Excel导入剩余产品到ciciERP数据库

跳过已导入的前17个产品(SP-01到SP-17)
导入剩余的119个产品(SP-18到SP-136)
"""

import sqlite3
import openpyxl
from datetime import datetime

# 配置
EXCEL_PATH = '/home/wxy/.xiaozhi/files/file_1773984868713_a5025b1a_Foreign_Trade_Management_Template.xlsx.xlsx'
DB_PATH = '/home/wxy/data/ciciERP/cicierp.db'
SHEET_INDEX = 2  # 第三个sheet
HEADER_ROW = 3  # 表头在第3行
DATA_START_ROW = 4  # 数据从第4行开始
SKIP_COUNT = 17  # 跳过前17个已导入的产品
EXCHANGE_RATE = 6.81

# 列索引 (1-based)
COL_PRODUCT_NAME = 1
COL_MODEL = 2
COL_COST_RMB = 3
COL_COST_USD = 4
COL_PROFIT_MARGIN = 5
COL_ALIBABA_FEE = 6
COL_SELLING_PRICE = 7
COL_STOCK_QTY = 8
COL_SIZE = 9
COL_WEIGHT = 10
COL_NOTES = 11
COL_SUPPLIER = 12


def main():
    print("=== ciciERP 产品导入工具 ===\n")

    # 1. 加载Excel
    print(f"加载Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb.worksheets[SHEET_INDEX]
    print(f"Sheet: {sheet.title}")

    # 2. 连接数据库
    print(f"连接数据库: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 3. 获取已存在的产品编号
    cursor.execute("SELECT product_code FROM products WHERE deleted_at IS NULL")
    existing_codes = set(row[0] for row in cursor.fetchall())
    print(f"数据库已有 {len(existing_codes)} 个产品\n")

    # 4. 统计
    stats = {
        'total': 0,
        'success': 0,
        'skipped': 0,
        'failed': 0,
        'errors': []
    }

    # 5. 读取并导入产品
    # 找到所有有效产品行
    valid_rows = []
    for row_num in range(DATA_START_ROW, sheet.max_row + 1):
        model = sheet.cell(row=row_num, column=COL_MODEL).value
        if model and str(model).strip():
            valid_rows.append(row_num)

    print(f"Excel共有 {len(valid_rows)} 个产品")
    print(f"跳过前 {SKIP_COUNT} 个已导入产品\n")

    # 跳过前17个
    rows_to_import = valid_rows[SKIP_COUNT:]
    print(f"待导入 {len(rows_to_import)} 个产品\n")

    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

    for idx, row_num in enumerate(rows_to_import):
        # 读取数据
        product_name = sheet.cell(row=row_num, column=COL_PRODUCT_NAME).value
        model = str(sheet.cell(row=row_num, column=COL_MODEL).value).strip()
        cost_rmb = sheet.cell(row=row_num, column=COL_COST_RMB).value or 0
        cost_usd = sheet.cell(row=row_num, column=COL_COST_USD).value or 0
        profit_margin = sheet.cell(row=row_num, column=COL_PROFIT_MARGIN).value or 0
        alibaba_fee = sheet.cell(row=row_num, column=COL_ALIBABA_FEE).value
        selling_price = sheet.cell(row=row_num, column=COL_SELLING_PRICE).value
        stock_qty = sheet.cell(row=row_num, column=COL_STOCK_QTY).value
        size = sheet.cell(row=row_num, column=COL_SIZE).value
        weight = sheet.cell(row=row_num, column=COL_WEIGHT).value
        notes = sheet.cell(row=row_num, column=COL_NOTES).value
        supplier = sheet.cell(row=row_num, column=COL_SUPPLIER).value

        # 转换数值
        try:
            cost_rmb = float(cost_rmb) if cost_rmb else 0
            cost_usd = float(cost_usd) if cost_usd else 0
            profit_margin = float(profit_margin) if profit_margin else 0
            alibaba_fee = float(alibaba_fee) if alibaba_fee else None
            selling_price = float(selling_price) if selling_price else None
            stock_qty = int(stock_qty) if stock_qty else None
            weight = float(weight) if weight else None
        except (ValueError, TypeError) as e:
            stats['failed'] += 1
            stats['errors'].append(f"{model}: 数值转换错误 - {e}")
            continue

        stats['total'] += 1
        print(f"[{idx+1}/{len(rows_to_import)}] {model}: {product_name[:30] if product_name else ''}")

        # 检查是否已存在
        if model in existing_codes:
            stats['skipped'] += 1
            print("  跳过: 产品已存在")
            continue

        try:
            # 开始事务
            cursor.execute("BEGIN TRANSACTION")

            # 插入products表
            cursor.execute("""
                INSERT INTO products (
                    product_code, name, status,
                    purchase_price, sale_price,
                    description, created_at, updated_at,
                    name_en, slug, category_id, brand_id,
                    weight, volume, description_en,
                    specifications, main_image, images, is_featured, is_new,
                    view_count, sales_count
                ) VALUES (?, ?, 1, ?, ?, ?, ?, ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, '{}', NULL, '[]', 0, 0, 0, 0)
            """, (
                model,
                str(product_name)[:255] if product_name else model,
                cost_rmb,  # purchase_price
                cost_rmb,  # sale_price (参考售价)
                str(notes)[:500] if notes else None,  # description
                now, now
            ))

            product_id = cursor.lastrowid

            # 插入product_costs表
            cursor.execute("""
                INSERT INTO product_costs (
                    product_id, supplier_id, cost_cny, cost_usd, currency,
                    exchange_rate, profit_margin, platform_fee_rate, platform_fee,
                    sale_price_usd, quantity, purchase_order_id, is_reference,
                    effective_date, notes, created_at, updated_at
                ) VALUES (?, NULL, ?, ?, 'CNY', ?, 0, 0.025, NULL, NULL, 1, NULL, 1, NULL, ?, ?, ?)
            """, (
                product_id,
                cost_rmb,
                cost_usd,
                EXCHANGE_RATE,
                str(notes)[:500] if notes else None,
                now, now
            ))

            # 插入product_prices表（如果有售价信息）
            if selling_price:
                sale_price_cny = selling_price * EXCHANGE_RATE
                cursor.execute("""
                    INSERT INTO product_prices (
                        product_id, platform, sale_price_cny, sale_price_usd, exchange_rate,
                        profit_margin, platform_fee_rate, platform_fee, is_reference,
                        effective_date, notes, created_at, updated_at
                    ) VALUES (?, 'alibaba', ?, ?, ?, ?, 0.025, ?, 1, NULL, '从成本清单导入', ?, ?)
                """, (
                    product_id,
                    sale_price_cny,
                    selling_price,
                    EXCHANGE_RATE,
                    profit_margin,
                    alibaba_fee,
                    now, now
                ))

            conn.commit()
            stats['success'] += 1
            print("  成功")

        except Exception as e:
            conn.rollback()
            stats['failed'] += 1
            stats['errors'].append(f"{model}: {str(e)}")
            print(f"  失败: {e}")

    # 6. 输出统计
    print("\n=== 导入完成 ===")
    print(f"总计: {stats['total']}")
    print(f"成功: {stats['success']}")
    print(f"跳过: {stats['skipped']}")
    print(f"失败: {stats['failed']}")

    if stats['errors']:
        print("\n失败详情:")
        for error in stats['errors']:
            print(f"  - {error}")

    # 7. 验证导入数量
    cursor.execute("SELECT COUNT(*) FROM products WHERE deleted_at IS NULL")
    final_count = cursor.fetchone()[0]
    print(f"\n数据库产品总数: {final_count}")

    conn.close()
    wb.close()

    return stats['success'], stats['failed'], final_count


if __name__ == '__main__':
    main()
